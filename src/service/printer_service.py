import csv
import openpyxl

from openpyxl.styles import Font, Alignment, Border, Side
from typing import List

from src.utils.utils import get_logger, float_format_pt_br, format_date_pt_br, round_trunc, manual_trunc
from src.models.asset import Asset
from src.models.wallet import Wallet
from src.service.config_service import ConfigService


class PrinterService():
    def __init__(self, config = None):
        self.config = config or ConfigService()

    def print_to_xlsx(self, output_file: str, wallet: Wallet):
        file_parts = output_file.rsplit(".", 1)

        if len(wallet.stocks) > 0:
            self.__default_xlsx_printing(file_parts, 'stocks', wallet.stocks, 0.0)

        if len(wallet.reits) > 0:
            self.__default_xlsx_printing(file_parts, 'reits', wallet.reits, wallet.get_total_value_reit())

    def print_to_csv(self, output_file: str, wallet: Wallet):
        file_parts = output_file.rsplit(".", 1)

        if len(wallet.stocks) > 0:
            self.__default_printing(file_parts, 'stocks', wallet.stocks)

        if len(wallet.reits) > 0:
            self.__default_printing(file_parts, 'reits', wallet.reits, wallet.get_total_value_reit())

        if len(wallet.others) > 0:
            self.__default_printing(file_parts, 'others', wallet.others)

    def __default_printing(self, file_parts, sufix, assets: List[Asset], total: float=1):

        assets = sorted(assets, key=lambda item: item.ticker, reverse=False)

        with open(f"{file_parts[0]}_{sufix}.{file_parts[1]}", mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow([
                'ATIVO',
                'PRIMEIRA COMPRA',
                'ULTIMA COMPRA',
                'QUANTIDADE',
                'MÉDIA',
                'PREÇO',
                'INVESTIDO',
                'POSIÇÃO',
                '% ALOCAÇÃO',
                'RENDIMENTOS'])

            for asset in assets:
                if asset.quantity == 0:
                    # name = asset.name if asset.code == '' else asset.code
                    # get_logger().debug(f"Asset {name} has 0 quantity")
                    # writer.writerow([name, 0,  0,  0, '', '', 0, 0, 0])
                    continue

                get_logger().debug(f"Printing {asset.ticker}")
                first_date, last_date = asset.get_buy_dates()
                asset_value = asset.price * asset.quantity
                asset_alocation = (asset_value / total) * 100

                writer.writerow([
                    asset.ticker,
                    format_date_pt_br(first_date),
                    format_date_pt_br(last_date),
                    int(asset.quantity),
                    float_format_pt_br(asset.average_price),
                    float_format_pt_br(asset.price),
                    float_format_pt_br(asset.get_invested_amount()),
                    float_format_pt_br(asset_value),
                    float_format_pt_br(asset_alocation),
                    float_format_pt_br(asset.get_income_amount()),
                ])

    def __default_xlsx_printing(self, file_parts, sufix, assets: List[Asset], total: float=1):
        """
        Exporta dados de ativos para um arquivo Excel (.xlsx)

        Args:
            file_parts (list): Partes do nome do arquivo
            assets (list): Lista de ativos
            total (float): Valor total dos investimentos
        """
        assets = sorted(assets, key=lambda item: item.ticker, reverse=False)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Carteira de Investimentos'
        headers = [
            'ATIVO',
            'PRIMEIRA COMPRA',
            'ULTIMA COMPRA',
            'QUANTIDADE',
            'MÉDIA',
            'PREÇO',
            'INVESTIDO',
            'POSIÇÃO',
            '% ALOCAÇÃO',
            'RENDIMENTOS'
        ]

        # Add stylish headers
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row, asset in enumerate(assets, start=2):
            asset_value = 0.0
            asset_alocation = 0.0
            first_date, last_date = asset.get_buy_dates()

            if asset.quantity > 0:
                asset_value = asset.price * asset.quantity
                asset_alocation = (asset_value / total) * 100

            sheet.cell(row=row, column=1, value=asset.ticker)
            sheet.cell(row=row, column=2, value=format_date_pt_br(first_date))
            sheet.cell(row=row, column=3, value=format_date_pt_br(last_date))
            sheet.cell(row=row, column=4, value=int(asset.quantity))
            sheet.cell(row=row, column=5, value=manual_trunc(asset.average_price))
            sheet.cell(row=row, column=6, value=manual_trunc(asset.price))
            sheet.cell(row=row, column=7, value=manual_trunc(asset.get_invested_amount()))
            sheet.cell(row=row, column=8, value=round_trunc(asset_value))
            sheet.cell(row=row, column=9, value=round_trunc(asset_alocation))
            sheet.cell(row=row, column=10, value=asset.get_income_amount())

        # Adjusting column width
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 6)
            sheet.column_dimensions[column].width = adjusted_width

        output_filename = f'{file_parts[0]}_{sufix}.xlsx'
        workbook.save(output_filename)
        get_logger().debug(f'Exported to {output_filename}')
